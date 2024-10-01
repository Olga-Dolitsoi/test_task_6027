import openpyxl

# Load the workbook and select the worksheet
wb = openpyxl.load_workbook('/home/olga/PycharmProjects/test_task_6027/Файл 2.xlsx')  # Replace with your file name
ws = wb.active  # or specify the sheet name: wb['SheetName']

# Loop through column C, starting from row 1
for row in range(1, ws.max_row + 1):
    c_cell = ws[f'C{row}']  # Cell in column C
    i_cell = ws[f'G{row}']  # Corresponding cell in column I

    # Get the number format from the cell in column C
    i_cell.value = c_cell.number_format.replace("\\", '').upper()

# Save the modified workbook
wb.save('/home/olga/PycharmProjects/test_task_6027/Файл 2 fix.xlsx')
